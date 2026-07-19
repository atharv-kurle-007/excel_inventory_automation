from openpyxl import load_workbook

# load workbook
workbook = load_workbook(r"excel file path")

# load current sheet
sheet = workbook.active

# store total inventory value
total_inventory_value = 0

# store low stock item
low_stock_items = []

# traverse through excel sheet
for row in range(2, sheet.max_row+1):
    # extract data 
    product = sheet[f"A{row}"].value
    quantity = sheet[f"B{row}"].value
    price = sheet[f"C{row}"].value

    print(product, quantity, price)

    # calculate total inventory value
    value = quantity * price
    total_inventory_value += value

    # find low stock item
    if quantity < 10:
        low_stock_items.append(product)

print(low_stock_items)

print("\nTotal Inventory Value:", total_inventory_value)

print("Low stock items:")

# print low stock items
for item in low_stock_items:
    print("-", item)

# create new sheet
report = workbook.create_sheet("Inventory Report")

# write data in excel
report["A1"] = "Total Inventory Value"
report["B1"] = total_inventory_value

report["A3"] = "Low Stock Items"

row = 4

for item in low_stock_items:
    report[f"A{row}"] = item
    row += 1

workbook.save(r"storing file name")