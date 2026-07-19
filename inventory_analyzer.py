from openpyxl import load_workbook
import smtplib
from email.message import EmailMessage

# load workbook
workbook = load_workbook(r"C:\Users\Shree\Desktop\PYTHON\PEPPERMINT ROBOTICS\inventory_automation\Inventory.xlsx")

# load current sheet
sheet = workbook.active

# store total inventory value
total_inventory_value = 0

# store low stock item
low_stock_items = []

# traverse through excel sheet
for row in range(2, sheet.max_row+1):
    # extract data 
    product = sheet[f"A{row}"].value
    quantity = sheet[f"B{row}"].value
    price = sheet[f"C{row}"].value

    print(product, quantity, price)

    # calculate total inventory value
    value = quantity * price
    total_inventory_value += value

    # find low stock item
    if quantity < 10:
        low_stock_items.append(product)

print(low_stock_items)

print("\nTotal Inventory Value:", total_inventory_value)

print("Low stock items:")

# print low stock items
for item in low_stock_items:
    print("-", item)

# create new sheet
report = workbook.create_sheet("Inventory Report")

# write data in excel
report["A1"] = "Total Inventory Value"
report["B1"] = total_inventory_value

report["A3"] = "Low Stock Items"

row = 4

for item in low_stock_items:
    report[f"A{row}"] = item
    row += 1

workbook.save(r"C:\Users\Shree\Desktop\PYTHON\PEPPERMINT ROBOTICS\inventory_automation\inventory_report.xlsx")

# connect to gmail smtp protocol
# server = smtplib.SMTP("smtp.gmail.com", 587)

# # secure connection
# server.starttls()

# # server login
# server.login("", "")

# # create mail object
# email = EmailMessage()

# # set email fields
# email["From"] = "pmmeme100@gmail.com"
# email["To"] = "akurle402@gmail.com"
# email["Subject"] = "Inventory Data Report"
# email.set_content("Hi, This is excel report of inventory data")

# # get excel file
# with open(r"C:\Users\Shree\Desktop\PYTHON\PEPPERMINT ROBOTICS\inventory_automation\inventory_report.xlsx", "rb") as file:
#     data = file.read()

# # add attachment
# email.add_attachment(
#     data, 
#     maintype="application",
#     subtype="vnd.ms-excel",
#     filename="inventory_report.xlsx"
# )

# # send the mail
# server.send_message(email)

# # close server
# server.quit()

# print("Send email report to Atharv")





